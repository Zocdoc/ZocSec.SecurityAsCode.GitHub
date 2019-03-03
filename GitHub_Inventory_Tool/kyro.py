# kyro.py || part of ZocSec.SecurityAsCode.GitHub
#
# A tool for extracting important informaiton about all repos in an organization.

# Authors:	Gary Tsai @garymalaysia
#            Jay Ball @veggiespam
#			
#

from github import Github
from github import enable_console_debug_logging
from github import GithubException
from time import sleep
import argparse
import datetime
import sys
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Side, Alignment, Font

# Prior running this script do the following
# execute "pip3 install -r requirements.txt"


if __name__ == "__main__":
	parser = argparse.ArgumentParser(description='This tool allow user to audit Enterprise Repos.')
	parser.add_argument("-t", "--token", required=True, help="GitHub User Private Token")
	parser.add_argument("-a",'--audit',action='store_true',help="Audit github Repo")
	parser.add_argument("-u",'--update',help="Import Topics to Repo")
	args = vars(parser.parse_args())
	#audit_github(args["token"]) # user token is required

	count = 0
	row = 1
	num_of_keys = 1
	num_of_topic = 0
	human = 0
	col = 1

	# using an access token
	user = Github(args["token"])

# use in the tile with month and year
	month_of_year = datetime.date.today().strftime("%Y-%m-%d")

	if args["audit"]:

	# openpyxl at work
		wb = Workbook() #Creating excel Spreadsheet and initializing title
		ws = wb.active
		ws.title = "Github Inventory"
		filepath = "Github_Audit_"+month_of_year+".xlsx" #Excel spreadsheet name
		wb.save(filepath)
		wb=load_workbook(filepath)
		wb.create_sheet('Repos Deploy keys') # create a second sheet within the same excel spreadsheet
		deploy_keys_sheet = wb["Repos Deploy keys"] 
		wb.create_sheet('Summary') # create a third sheet within the same excel spreadsheet
		summary_sheet = wb["Summary"] 



		print ("Gathering all repos.. And apologize for the bootleg progress bar =_=!")
		
	# Using authenticated user token to gather all repo from Zocdoc
		org = user.get_organization("Zocdoc")
		repos = org.get_repos()
		for rep in repos:
			count +=1
	# For the bootleg progress bar
		toolbar_width = count
		sys.stdout.write("[%s]" % (" " * toolbar_width))
		sys.stdout.flush()
		sys.stdout.write("\b" * (toolbar_width+1)) # return to start of line, after '['

	# Coloring for the cell in the excel sheet
		alignment=Alignment(wrap_text=True)
		skyFill = PatternFill(start_color='add8e6',end_color='add8e6',fill_type="solid")
		redFill = PatternFill(start_color='fa8072',end_color='fa8072',fill_type="solid")
		grassFill = PatternFill(start_color='C1FFC1',end_color='C1FFC1',fill_type="solid")

	# Writing Columns to first sheet in the excel sheet
		github_inventory=wb.active
		github_inventory['a1'] = 'node_id'
		github_inventory['b1'] = 'Name'
		github_inventory['c1'] = 'Description'
		github_inventory['d1'] = 'Archived'
		github_inventory['e1'] = 'Private'
		github_inventory['f1'] = 'Languages'
		github_inventory['g1'] = 'Teams'
		github_inventory['h1'] = 'Topics'
		github_inventory['i1'] = 'Topic_add'
		github_inventory['j1'] = 'Topic_del'
		github_inventory['k1'] = 'Last Comitter'
		github_inventory['l1'] = 'Last Comit Date'
		github_inventory.freeze_panes = 'A2'                # Freeze top row
		github_inventory.column_dimensions['A'].width = 15  # In characters, not pixels
		github_inventory.column_dimensions['B'].width = 30  # In characters, not pixels
		github_inventory.column_dimensions['C'].width = 30  # In characters, not pixels
		github_inventory.column_dimensions['D'].width = 10  # In characters, not pixels
		github_inventory.column_dimensions['E'].width = 10  # In characters, not pixels
		github_inventory.column_dimensions['F'].width = 40  # In characters, not pixels
		github_inventory.column_dimensions['G'].width = 30  # In characters, not pixels
		github_inventory.column_dimensions['H'].width = 30  # In characters, not pixels
		github_inventory.column_dimensions['I'].width = 25  # In characters, not pixels
		github_inventory.column_dimensions['J'].width = 25  # In characters, not pixels
		github_inventory.column_dimensions['K'].width = 25  # In characters, not pixels
		github_inventory.column_dimensions['L'].width = 20  # In characters, not pixels

	# Writing Columns to second sheet in the excel sheet
		deploy_keys_sheet['a1'] = 'Repository'
		deploy_keys_sheet['b1'] = 'Deploy Key ID'
		deploy_keys_sheet.freeze_panes = 'A2'               # Freeze top row
		deploy_keys_sheet.column_dimensions['A'].width = 30
		deploy_keys_sheet.column_dimensions['B'].width = 60

	# Writing Columns to third sheet in the excel sheet
		summary_sheet['a1'] = 'Total Repo without Team Topic'
		summary_sheet['b1'] = 'Percentage'
		summary_sheet['c1'] = 'Total Repo with last Human Commit'
		summary_sheet['d1'] = 'Percentage'
		summary_sheet.freeze_panes = 'A2'               # Freeze top row
		summary_sheet.column_dimensions['A'].width = 25
		summary_sheet.column_dimensions['B'].width = 10
		summary_sheet.column_dimensions['C'].width = 30
		summary_sheet.column_dimensions['D'].width = 10

	# Github API is doing work using for loop
		for repo in user.get_user().get_repos():
			row+=1
			sys.stdout.write("*")
			sys.stdout.flush()
			languages = ""
			for lang in repo.get_languages():
				languages = languages + lang + " "
			teamlist = ""
			for team in repo.get_teams():
				teamlist = teamlist + team.name + "/" + team.permission + " "
			topics = ""
			for topic in repo.get_topics():
				topics = topics + " "+ topic

	# Check to see is the Repo empty. if so, print message to cell
			try:
				for commits in repo.get_commits():
					if commits.committer :
						github_inventory['k%d' % row] = commits.committer.login
						if commits.committer.login.endswith("-zocdoc"):
							human += 1
					break
			except GithubException as e:
				github_inventory['k%d' % row] = e.args[1]['message']

	# Color the field to light blue id is even row number
			if row %2 == 0 :
				github_inventory['a%d' % row].fill = skyFill
				github_inventory['b%d' % row].fill = skyFill
				github_inventory['c%d' % row].fill = skyFill
				github_inventory['d%d' % row].fill = skyFill
				github_inventory['e%d' % row].fill = skyFill
				github_inventory['f%d' % row].fill = skyFill
				github_inventory['g%d' % row].fill = skyFill
				github_inventory['h%d' % row].fill = skyFill
				github_inventory['i%d' % row].fill = skyFill
				github_inventory['j%d' % row].fill = skyFill
				github_inventory['k%d' % row].fill = skyFill
				github_inventory['l%d' % row].fill = skyFill

	# Fetching data with Github API 
			github_inventory['a%d' % row] = repo.id
			github_inventory['b%d' % row] = repo.name
			github_inventory['c%d' % row] = repo.description

	# Highlight archived to green if true
			if repo.archived == True: 
				github_inventory['d%d' % row].fill = grassFill
				github_inventory['d%d' % row] = repo.archived
			else: 
				github_inventory['d%d' % row] = repo.archived

	# Highlight Private to red if False
			if repo.private == False:
				github_inventory['e%d' % row].fill = redFill
				github_inventory['e%d' % row] = repo.private
			else:
				github_inventory['e%d' % row] = repo.private
			github_inventory['f%d' % row] = languages
			github_inventory['g%d' % row] = teamlist

			github_inventory['l%d' % row] = repo.pushed_at

			# this adds the sort mechanisms to the table.
			github_inventory.auto_filter.ref = "A:L"
			github_inventory.auto_filter.add_sort_condition = "A:A"
			github_inventory.auto_filter.add_sort_condition = "B:B"
			github_inventory.auto_filter.add_sort_condition = "C:C"
			github_inventory.auto_filter.add_sort_condition = "D:D"
			github_inventory.auto_filter.add_sort_condition = "E:E"
			github_inventory.auto_filter.add_sort_condition = "F:F"
			github_inventory.auto_filter.add_sort_condition = "G:G"
			github_inventory.auto_filter.add_sort_condition = "H:H"
			github_inventory.auto_filter.add_sort_condition = "I:I"
			github_inventory.auto_filter.add_sort_condition = "J:J"
			github_inventory.auto_filter.add_sort_condition = "K:K"
			github_inventory.auto_filter.add_sort_condition = "L:L"


	# Finding Repo with missing Team Topic
			if topics.find("team-") > 0:
				github_inventory['h%d' % row] = topics
			else:
				num_of_topic += 1
				github_inventory['h%d' % row].fill = redFill
				github_inventory['h%d' % row] = topics
				github_inventory['i%d' % row].fill = redFill
				

			
	# Fetching repo deploy keys with Github API 		
			for key in repo.get_keys():
				num_of_keys += 1
				deploy_keys_sheet['a%d' % num_of_keys] = repo.full_name
				deploy_keys_sheet['b%d' % num_of_keys] = str(key)

			sleep(0.3) # provide some seperation for each Github API call

		sys.stdout.write("\n") # Boot leg progress bar ending

# Calcualting formular to added to the third sheet for all missing team topic repo
		summary_sheet['a2'] = num_of_topic
		summary_sheet['b2'] = num_of_topic/count*100
		summary_sheet['b2'].fill = redFill
		summary_sheet['c2'] = human
		summary_sheet['d2'] = human/count*100
		summary_sheet['d2'].fill = skyFill

		wb.save(filepath) # Saving the excel file

# if -u flag is used, execute below script
	if args["update"]:
		wb = load_workbook(filename=args["update"],read_only=True)
		read_github_inventory = wb['Github Inventory']
		github_inventory_max_row = read_github_inventory.max_row
		for columns in range(1,read_github_inventory.max_column+1):

			# search for column "Topic_add"
			if read_github_inventory['%s1'% get_column_letter(columns)].value == "Topic_add":
				for row in range(2,github_inventory_max_row+1):
					if read_github_inventory['%s%d' % (get_column_letter(columns),row)].value: #if there is a value in the cell of column "Topic_add"
						topic_add = read_github_inventory['%s%d' % (get_column_letter(columns),row)].value.strip().split(" ") #remove leading and trailing spaces and split each word into a List

						# search for column "Topics"
						while read_github_inventory['%s1'% get_column_letter(col)].value != "Topics":
							col += 1 #iterate from column 1 on to look for "Topics"

						else: # if "Topics" is found, do next
							if read_github_inventory['%s%d' % (get_column_letter(col),row)].value:#look for data in the cell under the column "Topics"
								orginal_topic = read_github_inventory['%s%d' % (get_column_letter(col),row)].value.strip().split(" ")
								orginal_topic.extend(topic_add)# extending a list with data in "Topic_add"
								repo_name = read_github_inventory['b%d' % (row)].value # look for repo name in column B
								repo_events = user.get_repo("Zocdoc/%s" % repo_name)
								print (repo_name,orginal_topic)
								repo_events.replace_topics(orginal_topic)# add the extended list to GitHub
							else:# if no data is found in "Topics"
								repo_name = read_github_inventory['b%d' % (row)].value 
								repo_events = user.get_repo("Zocdoc/%s" % repo_name)
								print (repo_name, topic_add)
								repo_events.replace_topics(topic_add) # add whatever in "topic_add" column to GitHub

			if read_github_inventory['%s1'% get_column_letter(columns)].value == "Topic_del":# search for column "Topic_del"
				for row in range(2,github_inventory_max_row+1):
					if read_github_inventory['%s%d' % (get_column_letter(columns),row)].value:#if there is a value in the cell of column "Topic_del"
						topic_del = read_github_inventory['%s%d' % (get_column_letter(columns),row)].value.strip().split(" ") #remove leading and trailing spaces and split each word into a List
						while read_github_inventory['%s1'% get_column_letter(col)].value != "Topics":
							col += 1
						else:# if "Topics" is found, do next
							if read_github_inventory['%s%d' % (get_column_letter(col),row)].value: #look for data in the cell under the column "Topics"
								orginal_topic = read_github_inventory['%s%d' % (get_column_letter(col),row)].value.strip().split(" ") #remove leading and trailing spaces and split each word into a List

								updated_topics = [] #create a new list
								for i in orginal_topic: # look for element in the list "orginal_topic"
									if i not in topic_del: # compare List "topic_del" against List "orginal_topic"
										updated_topics.append(i) # Append what is not in List "topic_del" to List "updated_topics"
								
								repo_name = read_github_inventory['b%d' % (row)].value # look for repo name in column B
								repo_events = user.get_repo("Zocdoc/%s" % repo_name)
								print (orginal_topic,topic_del,updated_topics )
								repo_events.replace_topics(updated_topics) # Update Github with List "updated_topics"
							else: # if no value is found in column "Topic_del",
								pass # DON'T WORRY ABOUT IT!
							







# vim ts=4:noexpandtab
